import io
import re
import unicodedata
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from datetime import date, datetime, timedelta
from pathlib import Path
from urllib.parse import quote
from fastapi import FastAPI, Form, Request, UploadFile, File
from fastapi.responses import HTMLResponse, JSONResponse, PlainTextResponse, RedirectResponse, StreamingResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pathlib import Path

from openpyxl import Workbook
from PIL import Image, ImageOps
import pytesseract
import pypdfium2

from app.database import SessionLocal, Empresa, Contato, Contrato, NF, Anexo

app = FastAPI(title="Mercado Central - Gestão de Contratos")

LOGIN_USER = "mercado"
LOGIN_PASSWORD = "mercado123"
REMEMBER_ME_DAYS = 30

BASE_DIR = Path(__file__).resolve().parent.parent
app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))


def is_authenticated(request: Request) -> bool:
    return request.cookies.get("mc_session") == "authenticated"


def set_auth_cookie(response: RedirectResponse, remember_me: bool = False):
    max_age = REMEMBER_ME_DAYS * 24 * 60 * 60 if remember_me else None
    response.set_cookie(
        key="mc_session",
        value="authenticated",
        httponly=True,
        samesite="lax",
        max_age=max_age,
    )


def set_remembered_user(response: RedirectResponse, username: str, remember_me: bool):
    max_age = REMEMBER_ME_DAYS * 24 * 60 * 60 if remember_me else None
    if remember_me:
        response.set_cookie(
            key="mc_username",
            value=username,
            httponly=False,
            samesite="lax",
            max_age=max_age,
        )
        response.set_cookie(
            key="mc_remember_me",
            value="1",
            httponly=False,
            samesite="lax",
            max_age=max_age,
        )
    else:
        response.delete_cookie("mc_username")
        response.delete_cookie("mc_remember_me")


def parse_date(value):
    if not value:
        return None
    value = str(value).strip()
    if not value:
        return None

    try:
        return date.fromisoformat(value)
    except ValueError:
        pass

    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue

    return None


def parse_money_to_float(value):
    if value is None:
        return 0.0

    text = str(value).strip()
    if not text:
        return 0.0

    normalized = text.replace("R$", "").replace(" ", "").replace("_", "")

    if "," in normalized and "." in normalized:
        if normalized.rfind(",") > normalized.rfind("."):
            normalized = normalized.replace(".", "").replace(",", ".")
        else:
            normalized = normalized.replace(",", "")
    elif "," in normalized:
        normalized = normalized.replace(".", "").replace(",", ".")
    elif "." in normalized:
        if not re.search(r"\.\d{1,2}$", normalized):
            normalized = normalized.replace(".", "")

    try:
        return float(Decimal(normalized))
    except (InvalidOperation, ValueError):
        return 0.0


def ocr_image_bytes(image_bytes: bytes) -> str:
    if not image_bytes:
        return ""

    try:
        image = Image.open(io.BytesIO(image_bytes))
        if image.mode not in ("RGB", "L"):
            image = image.convert("RGB")
        image = ImageOps.grayscale(image)
        image = ImageOps.autocontrast(image)
        text = pytesseract.image_to_string(image, config="--psm 6 --oem 3") or ""
        return text.strip()
    except Exception:
        return ""


def extract_scanned_pdf_text(pdf_bytes: bytes) -> str:
    if not pdf_bytes or not pdf_bytes.startswith(b"%PDF"):
        return ""

    try:
        pdf = pypdfium2.PdfDocument(io.BytesIO(pdf_bytes))
        pages_text = []
        for page_index in range(len(pdf)):
            page = pdf[page_index]
            image = page.render(scale=2).to_pil()
            text = pytesseract.image_to_string(
                image,
                config="--psm 6 --oem 3",
            ) or ""
            if text.strip():
                pages_text.append(text.strip())
        return "\n".join(pages_text)
    except Exception:
        return ""


def extract_pdf_text(pdf_bytes: bytes) -> str:
    if not pdf_bytes:
        return ""

    if pdf_bytes.startswith(b"%PDF"):
        try:
            from pypdf import PdfReader
        except Exception:
            return ""

        try:
            reader = PdfReader(io.BytesIO(pdf_bytes))
            pages = []
            for page in reader.pages:
                text = page.extract_text() or ""
                pages.append(text)
            joined = "\n".join(pages)
            if joined.strip():
                return joined
        except Exception:
            pass

        scanned_text = extract_scanned_pdf_text(pdf_bytes)
        if scanned_text.strip():
            return scanned_text

    if pdf_bytes.startswith((b"\x89PNG", b"\xff\xd8\xff", b"GIF87a", b"GIF89a", b"\x42\x4d")):
        return ocr_image_bytes(pdf_bytes)

    for encoding in ("utf-8", "latin-1"):
        try:
            text = pdf_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
        if text.strip():
            return text

    return ocr_image_bytes(pdf_bytes)


def decode_mojibake(value: str) -> str:
    if not value:
        return ""

    repaired = value
    replacements = {
        "Â°": "º",
        "Â": "",
        "Ã": "Ã",
        "Ã©": "é",
        "Ã¡": "á",
        "Ã£": "ã",
        "Ã´": "ô",
        "Ã§": "ç",
        "Ãº": "ú",
        "Ã–": "Ö",
        "Ã©": "é",
        "Ã‰": "É",
        "Ã³": "ó",
        "Ãµ": "õ",
        "Ã\xa9": "é",
        "Ã\x81": "Á",
        "Ã\x83": "Ã",
        "Ã\x87": "Ç",
        "Ã\x89": "É",
        "Ã\x93": "“",
        "Ã\x94": "”",
        "Ã\x99": "™",
        "\x89": "é",
        "\x80": "",
        "�": "",
    }

    for wrong, correct in replacements.items():
        repaired = repaired.replace(wrong, correct)

    repaired = repaired.replace("\ufffd", "")
    repaired = repaired.replace("\x00", "")
    repaired = repaired.replace("\r", "\n")

    if "Ã" in repaired and "Â" in repaired:
        repaired = repaired.replace("Â", "")

    return repaired


def parse_invoice_from_text(text: str, filename: str = "") -> dict:
    raw_text = decode_mojibake(text or "")
    raw_text = raw_text.replace("�\x89", "é").replace("\x89", "é").replace("\x80", "")
    raw_text = raw_text.replace("\r", "\n")
    full_text = raw_text + "\n" + (decode_mojibake(filename or ""))
    result = {
        "empresa_nome": "",
        "numero_nf": "",
        "valor_nf": 0.0,
        "data_nf": "",
        "razao_nf": "",
    }

    def normalize_company_name(value: str) -> str:
        cleaned = (value or "").strip()
        cleaned = cleaned.replace("\u200b", "").replace("\x00", "")
        cleaned = re.sub(r"^(?:RAZ[AÃ]O\s+SOCIAL|RAZAO\s+SOCIAL|PRESTADOR\s*/\s*FORNECEDOR|EMITENTE\s*DA\s*NFS-E|NATUREZA\s+DA\s+OPERA[ÇC]AO|NATUREZA|OPERACAO|NOME\s*/\s*NOME\s+EMPRESARIAL|NOME\s+EMPRESARIAL|NOME\s+FANTASIA|NOME|EMPRESA|MUNIC[ÍI]PIO|CIDADE|UF|ESTADO|ENDERE[ÇC]O|BAIRRO|CEP|PA[ÍI]S)\s*[:\-]?\s*", "", cleaned, flags=re.I)
        cleaned = re.sub(r"\s+", " ", cleaned).strip(" -_")
        cleaned = re.sub(r"\s*(?:MUNIC[ÍI]PIO|CIDADE)\s*[:\-]?\s*[A-ZÀ-ŸÁÉÍÓÚÃÕÂÊÇ0-9 .'-]+(?:/\s*[A-Z]{2}|\s+[A-Z]{2})\s*$", "", cleaned, flags=re.I)
        cleaned = re.sub(r"\s*(?:UF|ESTADO)\s*[:\-]?\s*[A-Z]{2}\s*$", "", cleaned, flags=re.I)
        cleaned = re.sub(r"\s*[/\-]\s*[A-Z]{2}\s*$", "", cleaned, flags=re.I)
        cleaned = re.sub(r"\s+", " ", cleaned).strip(" -_")

        if re.search(r"\d", cleaned) and not re.search(r"[A-Za-zÀ-ÿ]", cleaned):
            return ""

        def strip_nfs_status_prefix(value_text: str) -> str:
            text = (value_text or "").strip()
            if not text:
                return ""

            ascii_text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
            pattern = r"^(?:SITUACAO\s*DA\s*NFS(?:-E)?|SITUACAO\s*DA\s*NFSE)\s*[:\-]?\s*(?:(?:EMITIDA|CANCELADA|PENDENTE|REJEITADA|DENEGADA|AUTORIZADA|NAO\s+EMITIDA|N[ÃA]O\s+EMITIDA)\s*[:\-]?\s*)?"
            stripped = re.sub(pattern, "", ascii_text, flags=re.I)
            if stripped != ascii_text:
                return stripped.strip()

            suffix_pattern = r"^(?:EMITIDA|CANCELADA|PENDENTE|REJEITADA|DENEGADA|AUTORIZADA|NAO\s+EMITIDA|N[ÃA]O\s+EMITIDA)\s*[:\-]?\s*"
            stripped = re.sub(suffix_pattern, "", ascii_text, flags=re.I)
            if stripped != ascii_text:
                return stripped.strip()
            return text

        cleaned = strip_nfs_status_prefix(cleaned)
        cleaned = re.sub(r"\s+", " ", cleaned).strip(" -_")

        if re.search(r"^(?:COMPRA|VENDA|MERCADORIA|SERVI[ÇC]O|SERVICO|OPERA[ÇC]AO|NATUREZA|TOTAL|VALOR|DATA|NF|NOTA|PRESTADOR|FORNECEDOR|FINALIDADE|INDICADOR\s+MUNICIPAL(?:\s*\([^)]*\))?|NFS(?:-E)?\s*(?:GERADA|REGULAR)|NOME\s*/\s*NOME\s+EMPRESARIAL|NOME\s+EMPRESARIAL|MUNIC[ÍI]PIO|CIDADE|UF|ESTADO|ENDERE[ÇC]O|BAIRRO|CEP|PA[ÍI]S|SITUA[ÇC]A[O0]?\s*DA\s*NFS(?:-E)?|EMITIDA|CANCELADA|PENDENTE|REJEITADA|DENEGADA|AUTORIZADA|NAO\s+EMITIDA|N[ÃA]O\s+EMITIDA|CNPJ\s*/\s*CPF\s*/\s*NIF|CNPJ|CPF|NIF|TELEFONE)", cleaned, flags=re.I):
            return ""

        if re.fullmatch(r"[\d\./\-()]+", cleaned):
            return ""

        if re.search(r"^(?:MUNIC[ÍI]PIO|CIDADE)\s*[:\-]?\s*[A-ZÀ-Ÿ\s]+(?:/\s*[A-Z]{2}|\s+[A-Z]{2})$", cleaned, flags=re.I):
            return ""
        if re.search(r"^(?:UF|ESTADO)\s*[:\-]?\s*[A-Z]{2}$", cleaned, flags=re.I):
            return ""

        if re.search(r"^(?:SITUA[ÇC]A[O0]?\s*DA\s*NFS(?:-E)?(?:\s*[:\-]?\s*(?:EMITIDA|CANCELADA|PENDENTE|REJEITADA|DENEGADA|AUTORIZADA|NAO\s+EMITIDA|N[ÃA]O\s+EMITIDA))?|EMITIDA|CANCELADA|PENDENTE|REJEITADA|DENEGADA|AUTORIZADA|NAO\s+EMITIDA|N[ÃA]O\s+EMITIDA|FINALIDADE|INDICADOR\s+MUNICIPAL|NFS(?:-E)?\s*(?:GERADA|REGULAR))$", cleaned, flags=re.I):
            return ""

        if re.search(r"(?:^|\s)mercado\s+central(?:\s+app)?(?:$|\s)", cleaned, flags=re.I) or re.search(r"^empresa\s+demo$", cleaned, flags=re.I):
            return ""

        if re.fullmatch(r"[A-ZÀ-Ÿ][A-Za-zÀ-ÿ' .-]{0,80}", cleaned) and not re.search(r"(?:LTDA|Ltda|S\.A|SA|MEI|EIRELI|INDUSTRIA|COM[ÉE]RCIO|SERVI[ÇC]OS|SERVICOS|BENS|CONSULTORIA|MATERIAIS|IMPORTA[ÇC]AO)", cleaned, flags=re.I):
            if re.search(r"^(?:[A-ZÀ-Ÿ]+|[A-ZÀ-Ÿ][a-zà-ÿ]+)$", cleaned) and len(cleaned.split()) <= 2:
                if cleaned.lower() in {"jundiai", "sao paulo", "campinas", "santos", "osasco", "guarulhos", "curitiba", "rio", "rio de janeiro", "belo horizonte", "porto alegre", "salvador", "recife", "fortaleza", "brasilia", "goiania", "manaus", "belem", "sp", "rj", "mg", "pr", "sc", "rs", "df", "ce", "pe", "ba", "ma", "pa"}:
                    return ""

        return cleaned

    def is_valid_company_candidate(value: str) -> bool:
        cleaned = normalize_company_name(value)
        if not cleaned or re.fullmatch(r"\d{1,12}", cleaned):
            return False
        if re.search(r"^(?:N(?:U|Ú|U)MERO|COMPET[ÊE]NCIA|VALOR|TOTAL|DATA|NF|NOTA|MUNIC[ÍI]PIO|CIDADE|UF|ESTADO|R\$)", cleaned, flags=re.I):
            return False
        return len(cleaned.strip()) >= 3

    def set_company_field(field_name: str, candidate: str):
        cleaned = normalize_company_name(candidate)
        if not cleaned or not is_valid_company_candidate(cleaned):
            return

        if field_name == "empresa_nome":
            if not result["empresa_nome"]:
                result["empresa_nome"] = cleaned
            elif result["empresa_nome"] == result["razao_nf"] and result["razao_nf"] != cleaned:
                result["empresa_nome"] = cleaned
        elif field_name == "razao_nf":
            if not result["razao_nf"]:
                result["razao_nf"] = cleaned
            elif result["razao_nf"] == result["empresa_nome"] and result["empresa_nome"] != cleaned:
                result["razao_nf"] = cleaned

    is_nfs_context = bool(re.search(r"(?:DANFSe|NFS-e|NFS-E)", full_text, flags=re.I))

    if not is_nfs_context:
        explicit_company_patterns = [
            ("empresa_nome", r"(?:NOME\s+FANTASIA|NOME\s*/\s*NOME\s+EMPRESARIAL|NOME\s+EMPRESARIAL|EMPRESA)\s*[:\-]?\s*([A-Za-zÀ-ÿ0-9 .&/()\-]+)"),
            ("razao_nf", r"(?:RAZ[AÃ]O\s+SOCIAL|RAZAO\s+SOCIAL|DENOMINA[ÇC]A[O0]\s+SOCIAL)\s*[:\-]?\s*([A-Za-zÀ-ÿ0-9 .&/()\-]+)"),
        ]
        for field_name, pattern in explicit_company_patterns:
            match = re.search(pattern, full_text, flags=re.I)
            if match:
                set_company_field(field_name, match.group(1))
                if field_name == "empresa_nome" and not result["razao_nf"]:
                    result["razao_nf"] = result["empresa_nome"]
                if field_name == "razao_nf" and not result["empresa_nome"]:
                    result["empresa_nome"] = result["razao_nf"]
                break

        cleaned_lines = [line.strip() for line in re.split(r"[\n\r]+", full_text) if line.strip()]
        if not result["empresa_nome"]:
            candidate_lines = []
            for line in cleaned_lines:
                repair_line = line.replace("�\x89", "é").replace("\x89", "é").replace("\x80", "")
                repair_line = normalize_company_name(repair_line)
                if not repair_line or len(repair_line) < 6:
                    continue
                if re.search(r"(?:NF|NFE|NOTA|FATURA|DANFE|CNPJ|IE|CHAVE|SÉRIE|SERIE|TOTAL|VALOR|PAGAMENTO|EMISS|DATA|R\$|COMPRA|VENDA|MERCADORIA|SERVI[ÇC]O|SERVICO|OPERA[ÇC]AO|NATUREZA|MUNIC[ÍI]PIO|CIDADE|UF|ESTADO|ENDERE[ÇC]O|BAIRRO)", repair_line, flags=re.I):
                    continue
                if re.search(r"[A-Za-zÀ-ÿ]", repair_line) and not re.search(r"\d", repair_line):
                    candidate_lines.append(repair_line)

            if candidate_lines:
                result["empresa_nome"] = normalize_company_name(max(candidate_lines, key=len))

    file_name = Path(filename or "").stem or ""
    file_name_normalized = file_name.strip()
    file_name_has_nf_marker = bool(re.search(
        r"(?i)(?:^|[_\-\s])(?:NF|NFE|NOTA|FATURA)(?:[_\-\s]*E)?(?:[_\-\s]*\d{1,12})",
        file_name_normalized,
    ))

    if file_name:
        if file_name_has_nf_marker:
            match = re.search(r"(?i)([A-Za-zÀ-ÿ][A-Za-zÀ-ÿ0-9 .&-]{4,80})(?:\s+(?:NF|NFE|NOTA|FATURA|DANFE)[-\s]?[A-Z0-9/.-]+)", file_name)
            if match and not result["empresa_nome"]:
                result["empresa_nome"] = normalize_company_name(match.group(1))

    def normalize_nf_number(raw_value: str) -> str:
        digits = re.sub(r"\D", "", str(raw_value or ""))
        if not digits or set(digits) == {"0"}:
            return ""
        return digits.lstrip("0") or "0"

    if is_nfs_context:
        lines = [line.strip() for line in re.split(r"[\n\r]+", full_text) if line.strip()]

        for index, line in enumerate(lines):
            normalized_line = re.sub(r"\s+", " ", line).strip()

            if re.search(r"(?:N(?:U|Ú|U)MERO\s*DA\s*NFS(?:-E)?|N(?:U|Ú|U)MERO\s*DA\s*NF(?:-E)?)", normalized_line, flags=re.I):
                for next_index in range(index + 1, min(index + 4, len(lines))):
                    candidate = re.sub(r"\s+", " ", lines[next_index]).strip()
                    if not candidate:
                        continue
                    number_match = re.search(r"([0-9]{1,12})", candidate)
                    if number_match:
                        result["numero_nf"] = normalize_nf_number(number_match.group(1))
                        break

            if re.search(r"(?:COMPET[ÊE]NCIA\s*DA\s*NFS-E|DATA\s+E\s+HORA\s+DA\s+EMISS[ÃA]O\s*DA\s*NFS-E|DATA\s+DE\s+EMISS[ÃA]O\s*DA\s*NFS-E|COMPET[ÊE]NCIA)", normalized_line, flags=re.I):
                for next_index in range(index + 1, min(index + 4, len(lines))):
                    candidate = re.sub(r"\s+", " ", lines[next_index]).strip()
                    date_match = re.search(r"([0-9]{2}/[0-9]{2}/[0-9]{4})", candidate)
                    if date_match:
                        result["data_nf"] = date_match.group(1)
                        break

            razao_match = re.search(r"(?:RAZ[AÃ]O\s+SOCIAL|RAZAO\s+SOCIAL|DENOMINA[ÇC]A[O0]\s+SOCIAL)\s*[:\-]?\s*([A-Za-zÀ-ÿ0-9 .&/()\-]+)", normalized_line, flags=re.I)
            if razao_match:
                repaired_candidate = normalize_company_name(razao_match.group(1))
                if repaired_candidate and is_valid_company_candidate(repaired_candidate) and not result["razao_nf"]:
                    result["razao_nf"] = repaired_candidate

            if re.search(r"(?:RAZ[AÃ]O\s+SOCIAL|RAZAO\s+SOCIAL|DENOMINA[ÇC]A[O0]\s+SOCIAL)\s*[:\-]?", normalized_line, flags=re.I):
                for next_index in range(index + 1, min(index + 4, len(lines))):
                    candidate = re.sub(r"\s+", " ", lines[next_index]).strip()
                    if not candidate:
                        continue
                    if re.search(r"^(?:RAZ[AÃ]O\s+SOCIAL|RAZAO\s+SOCIAL|DENOMINA[ÇC]A[O0]\s+SOCIAL|N(?:U|Ú|U)MERO|COMPET[ÊE]NCIA|CNPJ|CPF|NIF|TELEFONE|ENDERE[ÇC]O|EMAIL|MUNIC[ÍI]PIO|UF|ESTADO)", candidate, flags=re.I):
                        continue
                    repaired_candidate = normalize_company_name(candidate)
                    if repaired_candidate and is_valid_company_candidate(repaired_candidate) and not result["razao_nf"]:
                        result["razao_nf"] = repaired_candidate
                        break

            if re.search(r"^(?:INDICADOR\s+MUNICIPAL(?:\s*\([^)]*\))?|FINALIDADE|NFS(?:-E)?\s*(?:GERADA|REGULAR)|SITUA[ÇC]A[O0]?\s*DA\s*NFS(?:-E)?|EMITIDA|CANCELADA|PENDENTE|REJEITADA|DENEGADA|AUTORIZADA|NAO\s+EMITIDA|N[ÃA]O\s+EMITIDA)\s*[:\-]?(?:\s*(?:EMITIDA|CANCELADA|PENDENTE|REJEITADA|DENEGADA|AUTORIZADA|NAO\s+EMITIDA|N[ÃA]O\s+EMITIDA))?$", normalized_line, flags=re.I):
                for next_index in range(index + 1, min(index + 5, len(lines))):
                    candidate = re.sub(r"\s+", " ", lines[next_index]).strip()
                    if not candidate:
                        continue
                    if re.search(r"^(?:INDICADOR\s+MUNICIPAL(?:\s*\([^)]*\))?|FINALIDADE|NFS(?:-E)?\s*(?:GERADA|REGULAR)|SITUA[ÇC]A[O0]?\s*DA\s*NFS(?:-E)?|EMITIDA|CANCELADA|PENDENTE|REJEITADA|DENEGADA|AUTORIZADA|NAO\s+EMITIDA|N[ÃA]O\s+EMITIDA|N(?:U|Ú|U)MERO|COMPET[ÊE]NCIA|VALOR|TOTAL|R\$|CNPJ|CPF|MUNIC[ÍI]PIO|UF|ESTADO)", candidate, flags=re.I):
                        continue
                    repaired_candidate = normalize_company_name(candidate)
                    if repaired_candidate and is_valid_company_candidate(repaired_candidate):
                        if not result["empresa_nome"]:
                            result["empresa_nome"] = repaired_candidate
                        if not result["razao_nf"]:
                            result["razao_nf"] = repaired_candidate
                        break
                continue

            status_label_match = re.search(r"^(?:SITUA[ÇC]A[O0]?\s*DA\s*NFS(?:-E)?|EMITIDA|CANCELADA|PENDENTE|REJEITADA|DENEGADA|AUTORIZADA|NAO\s+EMITIDA|N[ÃA]O\s+EMITIDA)\s*[:\-]?(?:\s*(?:EMITIDA|CANCELADA|PENDENTE|REJEITADA|DENEGADA|AUTORIZADA|NAO\s+EMITIDA|N[ÃA]O\s+EMITIDA))?\s*(.*)$", normalized_line, flags=re.I)
            if not status_label_match:
                ascii_line = unicodedata.normalize("NFKD", normalized_line).encode("ascii", "ignore").decode("ascii")
                status_label_match = re.search(r"^(?:SITUACAO\s*DA\s*NFS(?:-E)?|SITUACAO\s*DA\s*NFSE|EMITIDA|CANCELADA|PENDENTE|REJEITADA|DENEGADA|AUTORIZADA|NAO\s+EMITIDA|N[AO]O\s+EMITIDA)\s*[:\-]?(?:\s*(?:EMITIDA|CANCELADA|PENDENTE|REJEITADA|DENEGADA|AUTORIZADA|NAO\s+EMITIDA|N[AO]O\s+EMITIDA))?\s*(.*)$", ascii_line, flags=re.I)

            if status_label_match:
                remaining = (status_label_match.group(1) or "").strip()
                if remaining:
                    repaired_candidate = normalize_company_name(remaining)
                    if repaired_candidate and is_valid_company_candidate(repaired_candidate):
                        if not result["empresa_nome"]:
                            result["empresa_nome"] = repaired_candidate
                        if not result["razao_nf"]:
                            result["razao_nf"] = repaired_candidate
                for next_index in range(index + 1, min(index + 4, len(lines))):
                    candidate = re.sub(r"\s+", " ", lines[next_index]).strip()
                    if not candidate:
                        continue
                    if re.search(r"^(?:SITUA[ÇC]A[O0]?\s*DA\s*NFS(?:-E)?|EMITIDA|CANCELADA|PENDENTE|REJEITADA|DENEGADA|AUTORIZADA|NAO\s+EMITIDA|N[ÃA]O\s+EMITIDA|N(?:U|Ú|U)MERO|COMPET[ÊE]NCIA|VALOR|TOTAL|R\$|CNPJ|CPF|MUNIC[ÍI]PIO|UF|ESTADO)", candidate, flags=re.I):
                        continue
                    repaired_candidate = normalize_company_name(candidate)
                    if repaired_candidate and is_valid_company_candidate(repaired_candidate):
                        if not result["empresa_nome"]:
                            result["empresa_nome"] = repaired_candidate
                        if not result["razao_nf"]:
                            result["razao_nf"] = repaired_candidate
                        break

            nome_empresa_match = re.search(r"(?:PRESTADOR\s*/\s*FORNECEDOR|EMITENTE\s*DA\s*NFS-E|NOME\s*/\s*NOME\s+EMPRESARIAL|NOME\s+EMPRESARIAL|NOME\s+FANTASIA)\s*[:\-]?\s*([A-Za-zÀ-ÿ0-9 .&/()\-]+)", normalized_line, flags=re.I)
            if nome_empresa_match:
                repaired_candidate = normalize_company_name(nome_empresa_match.group(1))
                if repaired_candidate and is_valid_company_candidate(repaired_candidate) and not result["empresa_nome"]:
                    result["empresa_nome"] = repaired_candidate

            if re.search(r"(?:PRESTADOR\s*/\s*FORNECEDOR|EMITENTE\s*DA\s*NFS-E|NOME\s*/\s*NOME\s+EMPRESARIAL|NOME\s+EMPRESARIAL|NOME\s+FANTASIA)\s*[:\-]?", normalized_line, flags=re.I):
                for next_index in range(index + 1, min(index + 8, len(lines))):
                    candidate = re.sub(r"\s+", " ", lines[next_index]).strip()
                    if not candidate:
                        continue
                    if re.search(r"^(?:PRESTADOR|FORNECEDOR|EMITENTE|NOME\s*/\s*NOME\s+EMPRESARIAL|NOME\s+EMPRESARIAL|NOME\s+FANTASIA|N(?:U|Ú|U)MERO|COMPET[ÊE]NCIA|CNPJ|CPF|NIF|TELEFONE|ENDERE[ÇC]O|EMAIL)", candidate, flags=re.I):
                        continue
                    repaired_candidate = normalize_company_name(candidate)
                    if repaired_candidate and is_valid_company_candidate(repaired_candidate) and not result["empresa_nome"]:
                        result["empresa_nome"] = repaired_candidate
                        break

            if re.search(r"(?:VALOR\s*L[ÍI]QUIDO\s*NFS(?:-E)?\s*(?:\+\s*IBS/CBS|\+\s*IBS|\+\s*CBS)?|VALOR\s*DA\s*OPERA(?:[ÇC]?[ÃA]?)?O/SERVI(?:[ÇC]?)?O|VALOR\s*TOTAL\s*DA\s*NFS-E|VALOR\s*L[ÍI]QUIDO\s*DA\s*NFS-E)", normalized_line, flags=re.I):
                for next_index in range(index + 1, min(index + 3, len(lines))):
                    candidate = re.sub(r"\s+", " ", lines[next_index]).strip()
                    amount_match = re.search(r"(?:R\$\s*)?([0-9]{1,3}(?:\.[0-9]{3})*(?:[.,][0-9]{2})|[0-9]+(?:[.,][0-9]{2}))", candidate, flags=re.I)
                    if amount_match:
                        result["valor_nf"] = parse_money_to_float(amount_match.group(1))
                        break

        if not result["empresa_nome"] and result["razao_nf"]:
            result["empresa_nome"] = result["razao_nf"]
        if not result["razao_nf"] and result["empresa_nome"]:
            result["razao_nf"] = result["empresa_nome"]

    if not is_nfs_context:
        numero_patterns = [
            r"(?:NF(?:[- ]?E)?|NFE|NOTA\s*FISCAL|N[º°]|NÚMERO|NUMERO)\s*[:\-]*\s*(?:N[º°]\s*)?([0-9]{1,12})",
            r"(?:NF(?:[- ]?E)?|NFE|NOTA\s*FISCAL|N[º°]|NÚMERO|NUMERO)[^\d]{0,15}([0-9]{1,12})",
        ]
        for pattern in numero_patterns:
            numero_match = re.search(pattern, full_text, flags=re.I)
            if numero_match:
                candidate_number = normalize_nf_number(numero_match.group(1))
                if candidate_number:
                    result["numero_nf"] = candidate_number
                    break

        if not result["numero_nf"]:
            for line in re.split(r"[\n\r]+", full_text):
                if re.search(r"(?:NF|NFE|NOTA FISCAL|N[º°]|NÚMERO|NUMERO)", line, flags=re.I):
                    numbers = re.findall(r"\d{3,12}", line)
                    for number in numbers:
                        candidate_number = normalize_nf_number(number)
                        if candidate_number:
                            result["numero_nf"] = candidate_number
                            break
                    if result["numero_nf"]:
                        break

        if not result["numero_nf"] and file_name_has_nf_marker:
            for pattern in numero_patterns:
                filename_match = re.search(pattern, file_name, flags=re.I)
                if filename_match:
                    candidate_number = normalize_nf_number(filename_match.group(1))
                    if candidate_number:
                        result["numero_nf"] = candidate_number
                        break

    if not is_nfs_context:
        final_total_pattern = r"(?:valor\s*(?:total|liquido)(?!\s+(?:dos\s+produtos|de\s+produtos)\b)(?:\s*(?:da\s*)?(?:nota|nf))?|total\s*(?:da\s*)?(?:nota|nf)|total\s*geral)"
        product_total_pattern = r"(?:valor\s*(?:total|liquido)\s*(?:dos\s*produtos|de\s*produtos)|total\s*(?:dos\s*produtos|de\s*produtos)|valor\s*dos\s*produtos|subtotal)"

        explicit_total_match = None
        for line in re.split(r"[\n\r]+", full_text):
            normalized_line = line.strip()
            if not normalized_line:
                continue

            if re.search(final_total_pattern, normalized_line, flags=re.I):
                amount_match = re.search(r"(?:R\$\s*)?([0-9]{1,3}(?:\.[0-9]{3})*(?:[.,][0-9]{2})|[0-9]+(?:[.,][0-9]{2}))", normalized_line, flags=re.I)
                if amount_match:
                    explicit_total_match = amount_match
                    break

        if explicit_total_match:
            result["valor_nf"] = parse_money_to_float(explicit_total_match.group(1))
        else:
            product_total_match = None
            for line in re.split(r"[\n\r]+", full_text):
                normalized_line = line.strip()
                if not normalized_line:
                    continue
                if re.search(product_total_pattern, normalized_line, flags=re.I):
                    amount_match = re.search(r"(?:R\$\s*)?([0-9]{1,3}(?:\.[0-9]{3})*(?:[.,][0-9]{2})|[0-9]+(?:[.,][0-9]{2}))", normalized_line, flags=re.I)
                    if amount_match:
                        product_total_match = amount_match
                        break

            if product_total_match:
                result["valor_nf"] = parse_money_to_float(product_total_match.group(1))
            else:
                money_values = re.findall(r"(?:R\$\s*)?([0-9]{1,3}(?:\.[0-9]{3})*(?:[.,][0-9]{2})|[0-9]+(?:[.,][0-9]{2}))", full_text, flags=re.I)
                if money_values:
                    result["valor_nf"] = parse_money_to_float(money_values[-1])

    if result["valor_nf"] == 0 and file_name_has_nf_marker:
        filename_value_match = re.search(r"(?:R\$\s*)?(\d{1,3}(?:\.\d{3})*(?:[.,]\d{2})|\d+(?:[.,]\d{2}))", file_name)
        if filename_value_match:
            result["valor_nf"] = parse_money_to_float(filename_value_match.group(1))

    date_context_match = re.search(r"(?:data\s*(?:de\s*)?(?:emiss[ãa]o|recebimento|emissao)|emissao)[:\s-]*([0-9]{2}/[0-9]{2}/[0-9]{4}|[0-9]{2}-[0-9]{2}-[0-9]{4}|[0-9]{4}-[0-9]{2}-[0-9]{2})", full_text, flags=re.I)
    if date_context_match:
        result["data_nf"] = date_context_match.group(1)
    else:
        date_match = re.search(r"(\d{2}/\d{2}/\d{4}|\d{2}-\d{2}-\d{4}|\d{4}-\d{2}-\d{2})", full_text)
        if date_match:
            result["data_nf"] = date_match.group(1)

    if not result["data_nf"]:
        date_match = re.search(r"(\d{2}/\d{2}/\d{4}|\d{2}-\d{2}-\d{4}|\d{4}-\d{2}-\d{2})", file_name)
        if date_match:
            result["data_nf"] = date_match.group(1)

    result["empresa_nome"] = normalize_company_name(result["empresa_nome"])
    result["razao_nf"] = normalize_company_name(result["razao_nf"] or result["empresa_nome"])
    if not result["empresa_nome"] and result["razao_nf"]:
        result["empresa_nome"] = result["razao_nf"]
    if not result["razao_nf"] and result["empresa_nome"]:
        result["razao_nf"] = result["empresa_nome"]
    return result


def round_money(value):
    if value is None:
        return 0.0
    try:
        return float(Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
    except (InvalidOperation, ValueError, TypeError):
        return 0.0


def format_currency_br(value):
    try:
        numeric = Decimal(str(value or 0)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        formatted = f"{numeric:,.2f}"
        return formatted.replace(",", "_").replace(".", ",").replace("_", ".")
    except (InvalidOperation, ValueError, TypeError):
        return "0,00"


def normalize_cnpj(value: str) -> str:
    digits = re.sub(r"\D", "", str(value or ""))
    return digits[:14]


def is_valid_cnpj(value: str) -> bool:
    cnpj = normalize_cnpj(value)
    if len(cnpj) != 14 or cnpj == cnpj[0] * 14:
        return False

    def calc_digit(digits: str, weight: list[int]) -> str:
        total = sum(int(digit) * factor for digit, factor in zip(digits, weight))
        remainder = total % 11
        return "0" if remainder < 2 else str(11 - remainder)

    digits_base = cnpj[:12]
    first = calc_digit(digits_base, list(range(5, 1, -1)) + list(range(9, 1, -1)))
    second = calc_digit(digits_base + first, list(range(6, 1, -1)) + list(range(9, 1, -1)))
    return cnpj == digits_base + first + second


def format_cnpj(value: str) -> str:
    digits = normalize_cnpj(value)
    if len(digits) != 14:
        return digits
    return f"{digits[:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:14]}"


@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    if is_authenticated(request):
        return RedirectResponse(url="/", status_code=307)
    saved_username = request.cookies.get("mc_username", "").strip()
    remember_me = request.cookies.get("mc_remember_me") == "1"
    return templates.TemplateResponse(
        "login.html",
        {"request": request, "saved_username": saved_username, "remember_me": remember_me},
    )


@app.post("/login")
def login_submit(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    remember_me: str = Form("0"),
):
    remember_selected = remember_me == "1"
    if username == LOGIN_USER and password == LOGIN_PASSWORD:
        response = RedirectResponse(url="/", status_code=303)
        set_auth_cookie(response, remember_me=remember_selected)
        set_remembered_user(response, username, remember_selected)
        return response

    return templates.TemplateResponse(
        "login.html",
        {
            "request": request,
            "error": "Credenciais inválidas. Tente novamente.",
            "saved_username": username,
            "remember_me": remember_selected,
        },
    )


@app.get("/logout")
def logout():
    response = RedirectResponse(url="/login", status_code=307)
    response.delete_cookie("mc_session")
    response.delete_cookie("mc_username")
    response.delete_cookie("mc_remember_me")
    return response


@app.get("/", response_class=HTMLResponse)
def index(request: Request):
    if not is_authenticated(request):
        return RedirectResponse(url="/login", status_code=307)

    db = SessionLocal()
    empresas = db.query(Empresa).all()
    contatos = db.query(Contato).all()
    contratos = db.query(Contrato).all()
    nfs = db.query(NF).all()

    hoje = datetime.utcnow().date()
    inicio_dia = datetime.combine(hoje, datetime.min.time())
    fim_dia = inicio_dia + timedelta(days=1)

    resumo_dia = {
        "empresas": db.query(Empresa).filter(Empresa.data_cadastro >= inicio_dia, Empresa.data_cadastro < fim_dia).count(),
        "contatos": db.query(Contato).filter(Contato.data_cadastro >= inicio_dia, Contato.data_cadastro < fim_dia).count(),
        "contratos": db.query(Contrato).filter(Contrato.data_cadastro >= inicio_dia, Contrato.data_cadastro < fim_dia).count(),
        "nfs": db.query(NF).filter(NF.data_cadastro >= inicio_dia, NF.data_cadastro < fim_dia).count(),
    }
    db.close()

    success_message = request.query_params.get("msg")
    error_message = request.query_params.get("error")

    return templates.TemplateResponse(
        "index.html",
        {
            "request": request,
            "empresas": empresas,
            "contatos": contatos,
            "contratos": contratos,
            "nfs": nfs,
            "success_message": success_message,
            "error_message": error_message,
            "resumo_dia": resumo_dia,
            "format_currency_br": format_currency_br,
        },
    )


@app.post("/empresas")
def add_empresa(
    nome: str = Form(...),
    cnpj: str = Form(""),
    tipo: str = Form(""),
    status: str = Form("ativo"),
    observacoes: str = Form(""),
):
    cnpj_normalizado = normalize_cnpj(cnpj)
    if not cnpj_normalizado or not is_valid_cnpj(cnpj_normalizado):
        mensagem = "CNPJ obrigatório e inválido. Use o formato 00.000.000/0000-00."
        return RedirectResponse(url=f"/?error={quote(mensagem)}", status_code=303)

    db = SessionLocal()
    empresa = Empresa(
        nome=nome,
        cnpj=format_cnpj(cnpj_normalizado),
        tipo=tipo,
        status=status,
        observacoes=observacoes,
    )
    db.add(empresa)
    db.commit()
    db.close()

    mensagem = f"Empresa {nome} cadastrada"
    return RedirectResponse(url=f"/?msg={quote(mensagem)}", status_code=303)


@app.post("/contatos")
def add_contato(
    empresa_id: str = Form(""),
    empresa_nome: str = Form(""),
    nome: str = Form(...),
    cargo: str = Form(""),
    email: str = Form(""),
    telefone: str = Form(""),
    observacoes: str = Form(""),
):
    db = SessionLocal()
    empresa_id_value = None
    empresa_id_text = (empresa_id or "").strip()
    if empresa_id_text:
        try:
            empresa_id_value = int(empresa_id_text)
        except ValueError:
            empresa_id_value = None

    empresa_nome = (empresa_nome or "").strip()
    if not empresa_id_value and empresa_nome:
        empresa = db.query(Empresa).filter(Empresa.nome == empresa_nome).first()
        if not empresa:
            empresa = Empresa(
                nome=empresa_nome,
                cnpj="",
                tipo="Cliente",
                status="ativo",
                observacoes="Empresa criada automaticamente no cadastro de contato",
            )
            db.add(empresa)
            db.commit()
        empresa_id_value = empresa.id

    if not empresa_id_value:
        db.close()
        mensagem = "Selecione uma empresa ou informe o nome de uma nova empresa."
        return RedirectResponse(url=f"/?error={quote(mensagem)}", status_code=303)

    contato = Contato(
        empresa_id=empresa_id_value,
        nome=nome,
        cargo=cargo,
        email=email,
        telefone=telefone,
        observacoes=observacoes,
    )
    db.add(contato)
    db.commit()
    db.close()

    mensagem = f"Contato {nome} cadastrado"
    return RedirectResponse(url=f"/?msg={quote(mensagem)}", status_code=303)


@app.post("/contratos")
def add_contrato(
    empresa_id: str = Form(""),
    empresa_nome: str = Form(""),
    descricao: str = Form(""),
    categoria: str = Form("fixo"),
    tipo_servico: str = Form(""),
    data_assinatura: str = Form(""),
    vigencia: str = Form(""),
    renovacao: str = Form(""),
    denuncias: str = Form(""),
    valor_atual: float = Form(0.0),
    responsavel: str = Form(""),
    status: str = Form("ativo"),
    observacoes: str = Form(""),
):
    descricao_final = descricao.strip() or "Contrato"
    empresa_nome = (empresa_nome or "").strip()

    db = SessionLocal()
    empresa_id_value = None
    empresa_id_text = (empresa_id or "").strip()
    if empresa_id_text:
        try:
            empresa_id_value = int(empresa_id_text)
        except ValueError:
            empresa_id_value = None

    if not empresa_id_value and empresa_nome:
        empresa = db.query(Empresa).filter(Empresa.nome == empresa_nome).first()
        if not empresa:
            empresa = Empresa(
                nome=empresa_nome,
                cnpj="",
                tipo="Cliente",
                status="ativo",
                observacoes="Empresa criada automaticamente no cadastro de contrato",
            )
            db.add(empresa)
            db.commit()
        empresa_id_value = empresa.id

    if not empresa_id_value:
        db.close()
        mensagem = "Selecione uma empresa ou informe o nome de uma nova empresa para este contrato."
        return RedirectResponse(url=f"/?error={quote(mensagem)}", status_code=303)

    contrato = Contrato(
        empresa_id=empresa_id_value,
        descricao=descricao_final,
        categoria=categoria,
        tipo_servico=tipo_servico,
        data_assinatura=parse_date(data_assinatura),
        vigencia=vigencia,
        renovacao=renovacao,
        denuncias=denuncias.strip() if denuncias else "",
        valor_atual=round_money(valor_atual),
        responsavel=responsavel,
        status=status,
        observacoes=observacoes,
    )
    db.add(contrato)
    db.commit()
    db.close()

    nome_contrato = tipo_servico.strip() or categoria.strip() or "Serviço"
    mensagem = f"{nome_contrato.title()} cadastrado"
    return RedirectResponse(url=f"/?msg={quote(mensagem)}", status_code=303)


@app.get("/nfs/upload-page", response_class=HTMLResponse)
def upload_nf_page(request: Request):
    if not is_authenticated(request):
        return RedirectResponse(url="/login", status_code=307)

    db = SessionLocal()
    contratos = db.query(Contrato).all()
    db.close()

    return templates.TemplateResponse(
        "upload_nf.html",
        {
            "request": request,
            "contratos": contratos,
        },
    )


@app.post("/nfs/preview")
def preview_nf_upload(request: Request, file: UploadFile = File(...)):
    if not is_authenticated(request):
        return JSONResponse({"error": "sessao_invalidada"}, status_code=401)

    if not file or not getattr(file, "filename", None):
        return JSONResponse({"error": "arquivo_nao_enviado"}, status_code=400)

    uploaded_bytes = file.file.read()
    text = extract_pdf_text(uploaded_bytes)
    parsed = parse_invoice_from_text(text, file.filename)
    return {
        "empresa_nome": parsed.get("empresa_nome", ""),
        "numero_nf": parsed.get("numero_nf", ""),
        "valor_nf": float(parsed.get("valor_nf", 0.0) or 0.0),
        "data_nf": parsed.get("data_nf", ""),
        "razao_nf": parsed.get("razao_nf", ""),
    }


@app.post("/nfs")
def add_nf(
    contrato_id: str = Form(""),
    empresa_nome: str = Form(""),
    categoria: str = Form(""),
    tipo_servico: str = Form(""),
    descricao_contrato: str = Form(""),
    razao_nf: str = Form(""),
    data_nf: str = Form(""),
    numero_nf: str = Form(""),
    valor_nf: float = Form(0.0),
    data_pagamento: str = Form(""),
    forma_pagamento: str = Form(""),
    descricao_item: str = Form(""),
    observacoes: str = Form(""),
    status_conferencia: str = Form("pendente"),
    file: UploadFile = File(None),
):
    db = SessionLocal()
    categoria = (categoria or "").strip() or "fixo"

    parsed = {}
    if file and getattr(file, "filename", None):
        uploaded_bytes = file.file.read()
        parsed = parse_invoice_from_text(extract_pdf_text(uploaded_bytes), file.filename)

    empresa_nome = (empresa_nome or parsed.get("empresa_nome") or "").strip()
    razao_nf = (razao_nf or parsed.get("razao_nf") or empresa_nome or "").strip()
    numero_nf = (numero_nf or parsed.get("numero_nf") or "").strip()
    valor_nf = round_money(valor_nf or parsed.get("valor_nf") or 0)
    data_nf = parsed.get("data_nf") or data_nf
    data_pagamento = data_pagamento or ""

    contrato_id_value = None
    contrato_id_text = (contrato_id or "").strip()
    if contrato_id_text:
        try:
            contrato_id_value = int(contrato_id_text)
        except ValueError:
            contrato_id_value = None

    if not contrato_id_value:
        empresa = None
        if empresa_nome:
            empresa = db.query(Empresa).filter(Empresa.nome == empresa_nome).first()
            if not empresa:
                empresa = Empresa(
                    nome=empresa_nome,
                    cnpj="",
                    tipo="Cliente",
                    status="ativo",
                    observacoes="Empresa criada automaticamente no cadastro de NF",
                )
                db.add(empresa)
                db.commit()

        if empresa is None:
            db.close()
            mensagem = "Informe a empresa para criar automaticamente o contrato da NF."
            return RedirectResponse(url=f"/?error={quote(mensagem)}", status_code=303)

        contrato = db.query(Contrato).filter(
            Contrato.empresa_id == empresa.id,
            Contrato.categoria == categoria,
            Contrato.tipo_servico == (tipo_servico or "")
        ).order_by(Contrato.id.desc()).first()

        if not contrato:
            contrato = Contrato(
                empresa_id=empresa.id,
                descricao=(descricao_contrato or f"Contrato de {tipo_servico or categoria}").strip() or "Contrato automático",
                categoria=categoria,
                tipo_servico=tipo_servico or "Serviço",
                valor_atual=round_money(valor_nf),
                status="ativo",
                observacoes="Contrato gerado automaticamente no cadastro de NF",
            )
            db.add(contrato)
            db.commit()

        contrato_id_value = contrato.id

    nf = NF(
        contrato_id=contrato_id_value,
        razao_nf=razao_nf or empresa_nome or "",
        data_nf=parse_date(data_nf),
        numero_nf=numero_nf,
        valor_nf=round_money(valor_nf),
        data_pagamento=parse_date(data_pagamento),
        forma_pagamento=forma_pagamento,
        descricao_item=descricao_item,
        observacoes=observacoes,
        status_conferencia=status_conferencia,
    )
    db.add(nf)
    db.commit()
    db.close()

    mensagem = f"NF {numero_nf or 'registrada'} cadastrada"
    return RedirectResponse(url=f"/?msg={quote(mensagem)}", status_code=303)


@app.post("/nfs/upload")
def upload_nf_file(
    request: Request,
    empresa_nome: str = Form(""),
    categoria: str = Form("fixo"),
    tipo_servico: str = Form(""),
    descricao_contrato: str = Form(""),
    file: UploadFile = File(...),
):
    if not is_authenticated(request):
        return RedirectResponse(url="/login", status_code=307)

    fake_form = {
        "empresa_nome": empresa_nome,
        "categoria": categoria,
        "tipo_servico": tipo_servico,
        "descricao_contrato": descricao_contrato,
    }
    form = {**fake_form}
    return add_nf(
        contrato_id=form.get("contrato_id", ""),
        empresa_nome=form.get("empresa_nome", ""),
        categoria=form.get("categoria", "fixo"),
        tipo_servico=form.get("tipo_servico", ""),
        descricao_contrato=form.get("descricao_contrato", ""),
        razao_nf="",
        data_nf="",
        numero_nf="",
        valor_nf=0.0,
        data_pagamento="",
        forma_pagamento="",
        descricao_item="",
        observacoes="",
        status_conferencia="pendente",
        file=file,
    )


def get_relatorio_data(categoria: str | None = None):
    db = SessionLocal()
    nfs = db.query(NF).all()
    filtradas = []
    for nf in nfs:
        contrato = db.query(Contrato).filter(Contrato.id == nf.contrato_id).first() if nf.contrato_id else None
        nf.contrato = contrato
        nf.anexos = db.query(Anexo).filter(Anexo.nf_id == nf.id).all()
        if categoria is None or (contrato and contrato.categoria == categoria):
            filtradas.append(nf)
    db.close()
    return filtradas


def get_nf_alerts(limit: int = 5):
    hoje = date.today()
    alertas = []
    for nf in get_relatorio_data():
        if nf.data_pagamento is None:
            continue
        dias_restantes = (nf.data_pagamento - hoje).days
        if dias_restantes < 0:
            continue
        if dias_restantes <= 30:
            alertas.append({
                "nf": nf,
                "dias_restantes": dias_restantes,
                "nivel": "alerta" if dias_restantes <= 7 else "aviso",
            })
    alertas.sort(key=lambda item: item["dias_restantes"])
    return alertas[:limit]


def get_nf_status_summary(nfs):
    hoje = date.today()
    summary = {"pagas": 0, "pendentes": 0, "vencidas": 0}
    for nf in nfs:
        status = (nf.status_conferencia or "pendente").lower()
        if status == "confirmada":
            summary["pagas"] += 1
            continue
        if nf.data_pagamento is not None:
            if nf.data_pagamento < hoje:
                summary["vencidas"] += 1
            else:
                summary["pendentes"] += 1
        else:
            summary["pendentes"] += 1
    return summary


@app.get("/relatorio")
def relatorio(request: Request):
    if not is_authenticated(request):
        return RedirectResponse(url="/login", status_code=307)

    nfs = get_relatorio_data()
    fixos = get_relatorio_data("fixo")
    pontuais = get_relatorio_data("pontual")
    alertas = get_nf_alerts()
    status_totals = get_nf_status_summary(nfs)
    status_fixos = get_nf_status_summary(fixos)
    status_pontuais = get_nf_status_summary(pontuais)
    return templates.TemplateResponse(
        "relatorio.html",
        {
            "request": request,
            "nfs": nfs,
            "alertas": alertas,
            "categoria": "todos",
            "titulo": "Relatório geral",
            "fixos": fixos,
            "pontuais": pontuais,
            "status_totals": status_totals,
            "status_fixos": status_fixos,
            "status_pontuais": status_pontuais,
            "format_currency_br": format_currency_br,
        },
    )


@app.get("/relatorio/fixo")
def relatorio_fixo(request: Request):
    if not is_authenticated(request):
        return RedirectResponse(url="/login", status_code=307)

    nfs = get_relatorio_data("fixo")
    fixos = get_relatorio_data("fixo")
    pontuais = get_relatorio_data("pontual")
    return templates.TemplateResponse(
        "relatorio.html",
        {
            "request": request,
            "nfs": nfs,
            "alertas": get_nf_alerts(),
            "categoria": "fixo",
            "titulo": "Contratos fixos",
            "fixos": fixos,
            "pontuais": pontuais,
            "status_totals": get_nf_status_summary(nfs),
            "status_fixos": get_nf_status_summary(fixos),
            "status_pontuais": get_nf_status_summary(pontuais),
            "format_currency_br": format_currency_br,
        },
    )


@app.get("/relatorio/pontual")
def relatorio_pontual(request: Request):
    if not is_authenticated(request):
        return RedirectResponse(url="/login", status_code=307)

    nfs = get_relatorio_data("pontual")
    fixos = get_relatorio_data("fixo")
    pontuais = get_relatorio_data("pontual")
    return templates.TemplateResponse(
        "relatorio.html",
        {
            "request": request,
            "nfs": nfs,
            "alertas": get_nf_alerts(),
            "categoria": "pontual",
            "titulo": "Serviços pontuais",
            "fixos": fixos,
            "pontuais": pontuais,
            "status_totals": get_nf_status_summary(nfs),
            "status_fixos": get_nf_status_summary(fixos),
            "status_pontuais": get_nf_status_summary(pontuais),
            "format_currency_br": format_currency_br,
        },
    )


@app.post("/anexos")
def upload_anexo(
    request: Request,
    empresa_id: int = Form(...),
    contrato_id: int = Form(None),
    nf_id: int = Form(None),
    file: UploadFile = File(...),
):
    if not is_authenticated(request):
        return RedirectResponse(url="/login", status_code=307)

    conteudo = file.file.read()
    db = SessionLocal()
    anexo = Anexo(
        empresa_id=empresa_id,
        contrato_id=contrato_id or None,
        nf_id=nf_id or None,
        arquivo_nome=file.filename or "arquivo",
        mime_type=file.content_type or "application/octet-stream",
        tamanho=len(conteudo),
        conteudo=conteudo,
    )
    db.add(anexo)
    db.commit()
    db.close()

    mensagem = "Arquivo enviado"
    return RedirectResponse(url=f"/?msg={quote(mensagem)}", status_code=303)


@app.get("/anexos/{anexo_id}")
def download_anexo(anexo_id: int):
    db = SessionLocal()
    anexo = db.query(Anexo).filter(Anexo.id == anexo_id).first()
    db.close()
    if not anexo:
        return RedirectResponse(url="/", status_code=307)

    return Response(
        content=anexo.conteudo,
        media_type=anexo.mime_type or "application/octet-stream",
        headers={"Content-Disposition": f'attachment; filename="{anexo.arquivo_nome}"'},
    )


@app.get("/relatorio/excel")
def relatorio_excel(request: Request, categoria: str | None = None):
    if not is_authenticated(request):
        return RedirectResponse(url="/login", status_code=307)

    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    categoria = (categoria or "").strip().lower()
    if categoria in {"", "todos", "geral"}:
        categoria = "geral"
    elif categoria not in {"fixo", "pontual"}:
        categoria = "geral"

    db = SessionLocal()
    empresas = {empresa.id: empresa for empresa in db.query(Empresa).all()}
    contratos = db.query(Contrato).all()
    db.close()

    def get_last_nf_for_contrato(contrato_id: int):
        db = SessionLocal()
        nf = db.query(NF).filter(NF.contrato_id == contrato_id).order_by(NF.data_nf.desc().nullslast(), NF.id.desc()).first()
        db.close()
        return nf

    def get_rows_for_category(category_name: str):
        rows = []
        for contrato in contratos:
            if category_name == "geral":
                pass
            elif contrato.categoria != category_name:
                continue

            empresa = empresas.get(contrato.empresa_id)
            nf = get_last_nf_for_contrato(contrato.id)
            rows.append([
                empresa.nome if empresa else "Empresa não vinculada",
                contrato.descricao or "",
                contrato.categoria or "",
                contrato.tipo_servico or "",
                contrato.data_assinatura.strftime("%d/%m/%Y") if contrato.data_assinatura else "",
                contrato.vigencia or "",
                float(contrato.valor_atual) if contrato.valor_atual is not None else 0,
                contrato.responsavel or "",
                contrato.status or "",
                nf.numero_nf if nf else "",
                nf.data_nf.strftime("%d/%m/%Y") if nf and nf.data_nf else "",
                float(nf.valor_nf) if nf and nf.valor_nf is not None else 0,
                nf.data_pagamento.strftime("%d/%m/%Y") if nf and nf.data_pagamento else "",
                contrato.observacoes or "",
            ])
        return rows

    wb = Workbook()
    ws = wb.active
    ws.title = "Geral" if categoria == "geral" else ("Fixos" if categoria == "fixo" else "Pontuais")

    headers = [
        "Empresa",
        "Contrato",
        "Categoria",
        "Tipo de serviço",
        "Data de assinatura",
        "Vigência",
        "Valor atual",
        "Responsável",
        "Status",
        "Última NF",
        "Data NF",
        "Valor NF",
        "Data pagamento",
        "Observações",
    ]
    ws.append(headers)

    rows = get_rows_for_category(categoria if categoria in {"fixo", "pontual"} else "geral")
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="8B1D1D")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="E7D4CF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=14):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="center")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=7, max_col=7):
        for cell in row:
            cell.number_format = 'R$ #,##0.00'

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=11, max_col=11):
        for cell in row:
            if cell.value:
                cell.number_format = 'dd/mm/yyyy'

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=13, max_col=13):
        for cell in row:
            if cell.value:
                cell.number_format = 'dd/mm/yyyy'

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:N{ws.max_row}"
    ws.sheet_view.showGridLines = False
    widths = [24, 28, 14, 18, 16, 16, 14, 18, 14, 20, 12, 14, 14, 30]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    ws.print_title_rows = "1:1"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename_mapping = {
        "fixo": "relatorio_fixos_mercado_central.xlsx",
        "pontual": "relatorio_pontuais_mercado_central.xlsx",
        "geral": "relatorio_geral_mercado_central.xlsx",
    }
    response = StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = f"attachment; filename={filename_mapping[categoria]}"
    return response
