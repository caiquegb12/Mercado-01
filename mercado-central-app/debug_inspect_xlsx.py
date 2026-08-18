from openpyxl import load_workbook
from pathlib import Path

p = Path(r'C:\Users\caiqu\OneDrive\Desktop\Projeto Mercado Central\arquivos\CONTROLE PAGAMENTO CONTRATOS MENSAL (1).xlsx')
print('FILE_EXISTS', p.exists(), p)
wb = load_workbook(p, data_only=True)
print('SHEETS', wb.sheetnames)
ws = wb.active
print('MAX_ROW', ws.max_row, 'MAX_COL', ws.max_column)
for row in ws.iter_rows(min_row=1, max_row=min(25, ws.max_row), values_only=True):
    print(row)
