from openpyxl import load_workbook
from pathlib import Path

p = Path('C:/Users/caiqu/OneDrive/Desktop/Projeto Mercado Central/arquivos/CONTROLE PAGAMENTO CONTRATOS MENSAL (1).xlsx')
print('exists', p.exists())
wb = load_workbook(p, data_only=True)
print('sheets', wb.sheetnames)
ws = wb.active
print('rows', ws.max_row, 'cols', ws.max_column)
for index, row in enumerate(ws.iter_rows(min_row=1, max_row=min(30, ws.max_row), values_only=True), start=1):
    print(index, row)
