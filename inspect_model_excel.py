from openpyxl import load_workbook
from pathlib import Path

candidates = [
    Path(r'C:\Users\caiqu\OneDrive\Desktop\Projeto Mercado Central\arquivos\CONTROLE PAGAMENTO CONTRATOS MENSAL (1).xlsx'),
    Path(r'C:\Users\caiqu\OneDrive\Desktop\Projeto Mercado Central\arquivos\Controle de relação de pagamentos e NF\'s.xlsx'),
    Path(r'C:\Users\caiqu\OneDrive\Desktop\Projeto Mercado Central\arquivos\Controle de relação de pagamentos e NF\'s.xlsx'),
    Path(r'C:\Users\caiqu\OneDrive\Desktop\Projeto Mercado Central\arquivos\Controle de relação de pagamentos e NF\'s.xlsx'),
]
for p in candidates:
    print('TRY', p, 'EXISTS', p.exists())
    if p.exists():
        wb = load_workbook(p, data_only=True)
        print('SHEETS', wb.sheetnames)
        ws = wb.active
        print('ROWS', ws.max_row, 'COLS', ws.max_column)
        for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(30, ws.max_row), values_only=True), start=1):
            print(idx, row)
        print('---END---')
